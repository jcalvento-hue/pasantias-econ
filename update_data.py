import requests
import re
import json
from collections import Counter

FILE_ID = '1DpKK-YplfAcW9dWkeElaJQY_xWasnCfq'
DOWNLOAD_URL = f'https://drive.google.com/uc?export=download&id={FILE_ID}'

print('Downloading Excel...')
r = requests.get(DOWNLOAD_URL, allow_redirects=True)
r.raise_for_status()
content = r.content
print(f'Downloaded {len(content)} bytes')

with open('pasantias_raw', 'wb') as f:
    f.write(content)


def parse_spreadsheetml(data: bytes):
    """Excel 2003 XML (SpreadsheetML)"""
    from lxml import etree
    ns = 'urn:schemas-microsoft-com:office:spreadsheet'
    parser = etree.XMLParser(recover=True)
    root = etree.fromstring(data, parser)
    if root is None:
        raise ValueError('XML no parseable')
    sheets = root.findall(f'.//{{{ns}}}Worksheet')
    if not sheets:
        raise ValueError('No se encontró Worksheet en el XML')
    rows_xml = sheets[0].findall(f'.//{{{ns}}}Row')
    out = []
    for row in rows_xml:
        cells = row.findall(f'{{{ns}}}Cell')
        vals = []
        for c in cells:
            d = c.find(f'{{{ns}}}Data')
            vals.append(d.text if d is not None and d.text else '')
        out.append(vals)
    return out


def parse_xlsx(path: str):
    """Excel moderno (.xlsx, ZIP)"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(['' if v is None else str(v) for v in row])
    wb.close()
    return out


def parse_xls_binary(path: str):
    """Excel binario legacy (.xls, OLE)"""
    import xlrd
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    out = []
    for i in range(ws.nrows):
        vals = []
        for c in ws.row(i):
            if c.ctype == xlrd.XL_CELL_NUMBER and c.value == int(c.value):
                vals.append(str(int(c.value)))
            elif c.value is None:
                vals.append('')
            else:
                vals.append(str(c.value))
        out.append(vals)
    return out


# --- Detección de formato por magic bytes ---
head = content[:8]
if head.startswith(b'PK'):
    print('Formato detectado: XLSX (ZIP)')
    with open('pasantias.xlsx', 'wb') as f:
        f.write(content)
    all_rows = parse_xlsx('pasantias.xlsx')
elif head.startswith(b'\xd0\xcf\x11\xe0'):
    print('Formato detectado: XLS binario (OLE)')
    with open('pasantias.xls', 'wb') as f:
        f.write(content)
    all_rows = parse_xls_binary('pasantias.xls')
elif content.lstrip()[:1] == b'<' and b'<html' not in content[:500].lower():
    print('Formato detectado: SpreadsheetML (XML)')
    all_rows = parse_spreadsheetml(content)
else:
    print('ERROR: formato no reconocido. Primeros 500 bytes:')
    print(content[:500])
    raise SystemExit(
        'La descarga no es un Excel valido. Probablemente Drive devolvio '
        'una pagina HTML (permisos del archivo, o pagina de confirmacion).'
    )

if not all_rows:
    raise SystemExit('El archivo no contiene filas.')

header = all_rows[0]
records = [dict(zip(header, vals)) for vals in all_rows[1:]]
print(f'Parsed {len(records)} records')

# Build ANIOS_DATA
anios_raw = {}
for rec in records:
    f = rec.get('Fecha de inicio', '')
    if not f or len(f) < 4 or not f[:4].isdigit():
        continue
    anio = f[:4]
    if anio not in anios_raw:
        anios_raw[anio] = {'total': 0, 'activa': 0, 'finalizada': 0, 'montos': []}
    anios_raw[anio]['total'] += 1
    estado = rec.get('estado', '')
    if estado == 'activa':
        anios_raw[anio]['activa'] += 1
    elif estado == 'finalizada':
        anios_raw[anio]['finalizada'] += 1
    try:
        m = float(rec.get('Monto estimulo', '0') or 0)
        if m > 0:
            anios_raw[anio]['montos'].append(m)
    except Exception:
        pass

AD = {}
for anio in sorted(anios_raw.keys()):
    d = anios_raw[anio]
    avg = int(sum(d['montos']) / len(d['montos'])) if d['montos'] else 0
    AD[anio] = {'total': d['total'], 'activa': d['activa'], 'finalizada': d['finalizada'], 'montoAvg': avg}

# Build EMPRESAS
empresas = Counter(rec.get('Razon social', '').strip() for rec in records if rec.get('Razon social', '').strip())
top15 = empresas.most_common(15)
otras_total = sum(v for k, v in empresas.items() if k not in dict(top15))
EA = [{'name': k, 'total': v} for k, v in top15]
EA.append({'name': 'Otras empresas', 'total': otras_total})

# Build DBY (duration by year)
dur_labels = [1, 2, 3, 4, 5, 6, 9, 12]
DBY = {}
for rec in records:
    f = rec.get('Fecha de inicio', '')
    if not f or len(f) < 4 or not f[:4].isdigit():
        continue
    anio = f[:4]
    try:
        dur = int(float(rec.get('Duracion en meses', '0') or 0))
    except Exception:
        dur = 0
    if anio not in DBY:
        DBY[anio] = [0] * 8
    if dur in dur_labels:
        DBY[anio][dur_labels.index(dur)] += 1

# Serialize
ad_str = json.dumps(AD, ensure_ascii=False, separators=(',', ':'))
ea_str = json.dumps(EA, ensure_ascii=False, separators=(',', ':'))
dby_parts = [f'"{k}":{json.dumps(v, separators=(",", ":"))}' for k, v in sorted(DBY.items())]
dby_str = '{' + ','.join(dby_parts) + '}'
new_block = f'const AD={ad_str};\nconst EA={ea_str};\nconst DBY={dby_str};'

with open('index.html', 'r', encoding='utf-8') as f:
    html = f.read()
html, n_subs = re.subn(r'const AD=\{[\s\S]*?\};[\s]*const EA=[\s\S]*?\};[\s]*const DBY=\{[\s\S]*?\};', new_block, html)
if n_subs == 0:
    raise SystemExit('No se encontro el bloque const AD/EA/DBY en index.html — no se modifico nada.')
with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)
print('index.html updated successfully!')
